from typing import List
from timefold.solver.config import (SolverConfig, ScoreDirectorFactoryConfig,
                                    TerminationConfig, Duration)
from timefold.solver import SolverFactory, SolutionManager
from datetime import datetime
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict
from .domain import Employee, Shift, EmployeeSchedule
from .constraints import define_constraints

# Configure logging
logging.basicConfig(level=logging.INFO)
LOGGER = logging.getLogger('ShiftScheduler')


def main() -> None:
    department = "ALL"
    LOGGER.info(f"Running solver for {department}")
    do_work(department)


def do_work(department: str):

    solver_factory = SolverFactory.create(
        SolverConfig(
            solution_class=EmployeeSchedule,
            entity_class_list=[Shift],
            score_director_factory_config=ScoreDirectorFactoryConfig(
                constraint_provider_function=define_constraints
            ),
            termination_config=TerminationConfig(
                spent_limit=Duration(minutes=2),
                unimproved_spent_limit=Duration(minutes=1)
            )
        ))
    
    # Generate problem data
    problem = generate_demo_data(department)

    # Solve the problem
    solver = solver_factory.build_solver()
    LOGGER.info("Starting solver...")
    solution = solver.solve(problem)

    solution_manager = SolutionManager.create(solver_factory)
    #print(solution_manager.explain(solution).summary)
    score_analysis = solution_manager.analyze(solution)
    print(score_analysis.summary)

    for constraint_ref, constraint_analysis in score_analysis.constraint_map.items():
        #constraint_id = constraint_ref.constraint_id
        #score_per_constraint = constraint_analysis.score
        print(constraint_analysis.summary)

    # Output the solution
    LOGGER.info("Solution generated successfully.")

    create_shift_schedule_excel(solution.shifts, filename=f"shift_schedule_{department}.xlsx")


def create_shift_schedule_excel(shifts: list[Shift], filename: str):
    """
    Creates an Excel file with the shift schedule.
    """
    schedule = defaultdict(lambda: defaultdict(list))  # {date: {employee_name: [shifts]}}

    for shift in shifts:
        if shift.employee is None:
            continue
        shift_time = f"{shift.start.strftime('%H:%M')}-{shift.end.strftime('%H:%M')}"
        schedule[shift.start.date()][shift.employee.name].append(shift_time)

    sorted_dates = sorted(schedule.keys())
    unique_employees = sorted({shift.employee.name for shift in shifts if shift.employee})

    wb = Workbook()
    ws = wb.active
    ws.title = "Shift Schedule"

    ws.cell(row=1, column=1, value="Employee")
    for col, date in enumerate(sorted_dates, start=2):
        ws.cell(row=1, column=col, value=date.strftime("%d/%m/%Y"))

    for row, employee_name in enumerate(unique_employees, start=2):
        ws.cell(row=row, column=1, value=employee_name)
        for col, date in enumerate(sorted_dates, start=2):
            if schedule[date].get(employee_name):
                shifts_text = ", ".join(schedule[date][employee_name])
                ws.cell(row=row, column=col, value=shifts_text)

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    wb.save(filename)
    LOGGER.info(f"Excel file '{filename}' created successfully.")


def generate_demo_data(d) -> EmployeeSchedule:
    employees = []
   
    employees.append(Employee(name="A", skills=set(["C#", "Python", "SQL", "Linux"]), location="Sal", contract=8))
    employees.append(Employee(name="V", skills=set(["Python", "SQL"]), location="Ath", contract=8))
    employees.append(Employee(name="M", skills=set(["C#", "SQL"]), location="Ath", contract=8))
    employees.append(Employee(name="P", skills=set(["C#", "SQL", "Lua", "React"]), location="Ath", contract=8))
    employees.append(Employee(name="F", skills=set(["C#", "Python", "SQL", "Linux"]), location="Sal", contract=8))
    employees.append(Employee(name="L", skills=set(["C#", "SQL"]), location="Ath", contract=8))
    employees.append(Employee(name="I", skills=set(["C#", "SQL"]), location="Ath", contract=8))

    shifts = []

    # a full plus week of SQL requests...
    for i in range(1,31):
        shifts.append(Shift(id=f"1{i}".zfill(4), start=datetime(2024, 11, i, 9), end=datetime(2024, 11, i, 17), required_skill="SQL"))
        shifts.append(Shift(id=f"2{i}".zfill(4), start=datetime(2024, 11, i, 9), end=datetime(2024, 11, i, 17), required_skill="SQL"))
        shifts.append(Shift(id=f"3{i}".zfill(4), start=datetime(2024, 11, i, 9), end=datetime(2024, 11, i, 17), required_skill="SQL"))
        shifts.append(Shift(id=f"4{i}".zfill(4), start=datetime(2024, 11, i, 9), end=datetime(2024, 11, i, 17), required_skill="SQL"))
        shifts.append(Shift(id=f"5{i}".zfill(4), start=datetime(2024, 11, i, 9), end=datetime(2024, 11, i, 17), required_skill="SQL"))
        shifts.append(Shift(id=f"6{i}".zfill(4), start=datetime(2024, 11, i, 9), end=datetime(2024, 11, i, 17), required_skill="SQL"))
        shifts.append(Shift(id=f"7{i}".zfill(4), start=datetime(2024, 11, i, 9), end=datetime(2024, 11, i, 17), required_skill="SQL"))

    return EmployeeSchedule(employees=employees, shifts=shifts)


if __name__ == '__main__':
    main()
